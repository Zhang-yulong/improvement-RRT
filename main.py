from matplotlib import pyplot as plt
import re_informedRRT
import common
import openpyxl

show_animation = True


def main():
    print("Start rrt planning")

    obstacle1 = common.obstacle(3, 3, 3)
    obstacle2 = common.obstacle(5, 25, 5)
    obstacle3 = common.obstacle(9, 18, 4)
    obstacle4 = common.obstacle(12, 2, 2)
    obstacle5 = common.obstacle(16, 20, 3)
    obstacle6 = common.obstacle(22, 13, 5)
    obstacle7 = common.obstacle(27, 26, 2)

    obstacleList = [(obstacle1.x, obstacle1.y, obstacle1.r),
                    (obstacle2.x, obstacle2.y, obstacle2.r),
                    (obstacle3.x, obstacle3.y, obstacle3.r),
                    (obstacle4.x, obstacle4.y, obstacle4.r),
                    (obstacle5.x, obstacle5.y, obstacle5.r),
                    (obstacle6.x, obstacle6.y, obstacle6.r),
                    (obstacle7.x, obstacle7.y, obstacle7.r),
                    ]
    # Set params
    rrt = re_informedRRT.RRT(randArea=[-10, 40], obstacleList=obstacleList, maxIter=300)
    path = rrt.informed_rrt_star_planning(start=[0, 0], goal=[35, 33], animation=show_animation)
    print("Done!!")

    if show_animation and path:
        plt.show()

    #######################
    # 此处为测试新旧两种采样算法的（距离、时间、迭代次数），生成两种excel表格
    # 操作1：修改wb.save("旧采样方法.xlsx")
    # 操作2：进入re_informedRRT.py文件中的build_first_path()，
    #######################
    # wb = openpyxl.Workbook()
    # ws = wb.create_sheet("sheet1")
    # ws.cell(row=1, column=1).value = "路径长度"
    # ws.cell(row=1, column=2).value = "时间"
    # ws.cell(row=1, column=3).value = "迭代次数"
    # path = None
    # for i in range(1,101):
    #     path = rrt.informed_rrt_star_planning(start=[0, 0], goal=[35, 33], animation=show_animation)
    #     ws.cell(row=i + 1, column=1).value = path[1]
    #     print("1=", path[1])
    #     ws.cell(row=i + 1, column=2).value = path[2]
    #     print("2=", path[2])
    #     ws.cell(row=i + 1, column=3).value = path[3]
    #     print("3=", path[3])
    #     wb.save("新采样方法.xlsx")
    #     print("保存成功")


if __name__ == '__main__':
    main()
